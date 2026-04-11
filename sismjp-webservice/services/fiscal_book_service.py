"""
Gerador de Livro Fiscal em Excel (.xlsx).

Substitui o download do Livro Fiscal em PDF da prefeitura (que não está
disponível no webservice). Gera um arquivo Excel estruturado com todos os
campos das NFS-e, totais automáticos e formatação equivalente ao relatório
oficial da prefeitura de João Pessoa.

Dois tipos de livro:
  - Prestador: Serviços Prestados (notas emitidas)
  - Tomador:   Serviços Tomados (notas recebidas)
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.nfse_model import NfseData
from services.logging_service import log_error, log_info, log_warning

# ── Estilos ───────────────────────────────────────────────────────────────────

_AZUL_ESCURO = "003366"
_CINZA_CLARO = "E8E8E8"
_CINZA_TOTAL = "D0D0D0"

_FONTE_CABECALHO = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
_FONTE_DADO = Font(size=9, name="Calibri")
_FONTE_TOTAL = Font(bold=True, size=9, name="Calibri")
_FONTE_TITULO = Font(bold=True, size=12, name="Calibri")
_FONTE_SUBTITULO = Font(size=10, name="Calibri")

_FILL_CABECALHO = PatternFill("solid", fgColor=_AZUL_ESCURO)
_FILL_TOTAL = PatternFill("solid", fgColor=_CINZA_TOTAL)

_BORDA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_FMT_MOEDA = 'R$ #,##0.00'
_FMT_TEXTO = "@"

# ── Definição de colunas ──────────────────────────────────────────────────────
# (label, atributo_NfseData, largura, formato_numero)

_COLUNAS_PRESTADOR: list[tuple[str, str, int, str]] = [
    ("Nº NFS-e",          "numero",           14, _FMT_TEXTO),
    ("Data Emissão",      "data_emissao_fmt", 13, _FMT_TEXTO),
    ("Competência",       "competencia_fmt",  12, _FMT_TEXTO),
    ("Tomador",           "tomador_nome",     38, _FMT_TEXTO),
    ("CNPJ/CPF Tomador",  "tomador_cnpj",     18, _FMT_TEXTO),
    ("Cód. Serviço",      "codigo_servico",   12, _FMT_TEXTO),
    ("Descrição Serviço", "descricao_servico",45, _FMT_TEXTO),
    ("Valor Serviços",    "valor_servicos",   16, _FMT_MOEDA),
    ("Deduções",          "valor_deducoes",   13, _FMT_MOEDA),
    ("Base de Cálculo",   "base_calculo",     16, _FMT_MOEDA),
    ("Alíquota (%)",      "aliquota_pct",     12, "0.00"),
    ("Valor ISS",         "valor_iss",        13, _FMT_MOEDA),
    ("ISS Retido",        "iss_retido_str",   11, _FMT_TEXTO),
    ("PIS",               "valor_pis",        12, _FMT_MOEDA),
    ("COFINS",            "valor_cofins",     12, _FMT_MOEDA),
    ("INSS",              "valor_inss",       12, _FMT_MOEDA),
    ("IR",                "valor_ir",         12, _FMT_MOEDA),
    ("CSLL",              "valor_csll",       12, _FMT_MOEDA),
    ("Valor Líquido",     "valor_liquido",    16, _FMT_MOEDA),
    ("Situação",          "status_descricao", 12, _FMT_TEXTO),
]

_COLUNAS_TOMADOR: list[tuple[str, str, int, str]] = [
    ("Nº NFS-e",          "numero",           14, _FMT_TEXTO),
    ("Data Emissão",      "data_emissao_fmt", 13, _FMT_TEXTO),
    ("Competência",       "competencia_fmt",  12, _FMT_TEXTO),
    ("Prestador",         "prestador_nome",   38, _FMT_TEXTO),
    ("CNPJ Prestador",    "prestador_cnpj",   18, _FMT_TEXTO),
    ("Cód. Serviço",      "codigo_servico",   12, _FMT_TEXTO),
    ("Descrição Serviço", "descricao_servico",45, _FMT_TEXTO),
    ("Valor Serviços",    "valor_servicos",   16, _FMT_MOEDA),
    ("Deduções",          "valor_deducoes",   13, _FMT_MOEDA),
    ("Base de Cálculo",   "base_calculo",     16, _FMT_MOEDA),
    ("Alíquota (%)",      "aliquota_pct",     12, "0.00"),
    ("Valor ISS",         "valor_iss",        13, _FMT_MOEDA),
    ("ISS Retido",        "iss_retido_str",   11, _FMT_TEXTO),
    ("PIS",               "valor_pis",        12, _FMT_MOEDA),
    ("COFINS",            "valor_cofins",     12, _FMT_MOEDA),
    ("INSS",              "valor_inss",       12, _FMT_MOEDA),
    ("IR",                "valor_ir",         12, _FMT_MOEDA),
    ("CSLL",              "valor_csll",       12, _FMT_MOEDA),
    ("Valor Líquido",     "valor_liquido",    16, _FMT_MOEDA),
    ("Situação",          "status_descricao", 12, _FMT_TEXTO),
]

# Campos que recebem SUM na linha de totais
_CAMPOS_NUMERICOS = {
    "valor_servicos", "valor_deducoes", "base_calculo",
    "valor_iss", "valor_pis", "valor_cofins",
    "valor_inss", "valor_ir", "valor_csll", "valor_liquido",
}


# ── Função principal ──────────────────────────────────────────────────────────


def gerar_livro_fiscal(
    notas: list[NfseData],
    tipo: str,
    empresa_nome: str,
    empresa_codigo: str,
    competencia: str,
    comp_path: str,
    pasta_base: str,
) -> Path | None:
    """
    Gera o Livro Fiscal de um tipo (Prestador ou Tomador) como arquivo Excel.

    A estrutura do Excel espelha o relatório oficial da prefeitura de João Pessoa:
    - Linha 1: Título centralizado
    - Linha 2: Subtítulo com empresa e competência
    - Linha 3: Espaço
    - Linha 4: Cabeçalho das colunas (fundo azul, fonte branca)
    - Linhas 5+: Dados (uma por NFS-e)
    - Última linha: Totais com fórmulas SUM

    Args:
        notas: Lista de NfseData para o livro
        tipo: "Prestador" ou "Tomador"
        empresa_nome: Nome da empresa (usado no subtítulo e nome do arquivo)
        empresa_codigo: Código/IM (usado no nome do arquivo)
        competencia: Período de competência (ex: "04/2025")
        comp_path: Pasta de competência (ex: "042025")
        pasta_base: Pasta raiz de saída

    Returns:
        Path do arquivo gerado, ou None se não houver notas
    """
    if not notas:
        log_warning(
            empresa_nome,
            empresa_codigo,
            f"Livro Fiscal {tipo}",
            f"Nenhuma NFS-e encontrada. Livro Fiscal {tipo} não será gerado.",
        )
        return None

    colunas = _COLUNAS_PRESTADOR if tipo == "Prestador" else _COLUNAS_TOMADOR
    tipo_servico = "PRESTADOS" if tipo == "Prestador" else "TOMADOS"
    n_cols = len(colunas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Livro {tipo}"

    # ── Linha 1: Título ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titulo_cell = ws.cell(row=1, column=1)
    titulo_cell.value = f"PREFEITURA DE JOÃO PESSOA — LIVRO FISCAL DE SERVIÇOS {tipo_servico}"
    titulo_cell.font = _FONTE_TITULO
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Linha 2: Subtítulo ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = (
        f"Empresa: {empresa_nome}   |   Código: {empresa_codigo}   |   "
        f"Competência: {competencia}   |   Total de NFS-e: {len(notas)}"
    )
    sub_cell.font = _FONTE_SUBTITULO
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Linha 3: Espaço ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Linha 4: Cabeçalho ────────────────────────────────────────────────
    ROW_HEADER = 4
    for col_idx, (label, _, largura, _) in enumerate(colunas, start=1):
        cell = ws.cell(row=ROW_HEADER, column=col_idx, value=label)
        cell.font = _FONTE_CABECALHO
        cell.fill = _FILL_CABECALHO
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDA
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[ROW_HEADER].height = 32
    ws.freeze_panes = f"A{ROW_HEADER + 1}"

    # ── Linhas de dados ───────────────────────────────────────────────────
    ROW_DATA_START = ROW_HEADER + 1
    for row_offset, nota in enumerate(notas):
        row_num = ROW_DATA_START + row_offset
        for col_idx, (_, atributo, _, fmt) in enumerate(colunas, start=1):
            valor = getattr(nota, atributo, "")
            cell = ws.cell(row=row_num, column=col_idx, value=valor)
            cell.font = _FONTE_DADO
            cell.border = _BORDA
            if fmt == _FMT_MOEDA:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "0.00":
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=False)

    # ── Linha de totais ───────────────────────────────────────────────────
    ROW_TOTAL = ROW_DATA_START + len(notas)
    ws.cell(row=ROW_TOTAL, column=1, value="TOTAIS").font = _FONTE_TOTAL
    ws.cell(row=ROW_TOTAL, column=1).fill = _FILL_TOTAL
    ws.cell(row=ROW_TOTAL, column=1).alignment = Alignment(horizontal="left")

    for col_idx, (_, atributo, _, fmt) in enumerate(colunas, start=1):
        if atributo in _CAMPOS_NUMERICOS:
            col_letra = get_column_letter(col_idx)
            formula = f"=SUM({col_letra}{ROW_DATA_START}:{col_letra}{ROW_TOTAL - 1})"
            cell = ws.cell(row=ROW_TOTAL, column=col_idx, value=formula)
            cell.font = _FONTE_TOTAL
            cell.fill = _FILL_TOTAL
            cell.border = _BORDA
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
        elif col_idx > 1:
            cell = ws.cell(row=ROW_TOTAL, column=col_idx)
            cell.fill = _FILL_TOTAL
            cell.border = _BORDA

    # ── Auto-filtro ───────────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{ROW_HEADER}:{get_column_letter(n_cols)}{ROW_TOTAL - 1}"
    )

    # ── Salvar ────────────────────────────────────────────────────────────
    clean_nome = _limpar_nome(empresa_nome)
    output_dir = Path(pasta_base) / tipo / f"{empresa_codigo}-{clean_nome}" / comp_path
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{empresa_codigo}-{clean_nome}.xlsx"

    try:
        wb.save(str(output_path))
        log_info(
            empresa_nome,
            empresa_codigo,
            f"Livro Fiscal {tipo}",
            f"Gerado com {len(notas)} NFS-e → {output_path}",
        )
    except Exception as exc:
        log_error(
            empresa_nome, empresa_codigo, f"Livro Fiscal {tipo}",
            f"Erro ao salvar Excel: {exc}",
        )
        return None

    return output_path


def _limpar_nome(nome: str) -> str:
    """Remove caracteres inválidos para nomes de arquivo/pasta."""
    return re.sub(r'[<>:"/\\|?*]', "", nome).strip()
